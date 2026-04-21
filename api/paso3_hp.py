import pandas as pd
import os
from logger import get_logger

# Configurar logger
logger = get_logger("paso3_excel", log_dir="logs", log_file="paso3_excel.log")


def generar_excel(lista_datos, cancel_event, nombre_archivo="reporte_final.xlsx", callback_progreso=None, crear_excel=False ):
    """
    Genera un Excel Relacional con 5 pestañas.
    Flag crear_excel: Si es False, omite la generación pero retorna True para continuar el flujo.
    """
    total_items = len(lista_datos)
    
    # --- SKIP SI EL FLAG ESTÁ APAGADO ---
    if not crear_excel:
        logger.info("⏩ Generación de Excel desactivada (flag crear_excel=False). Saltando paso.")
        if callback_progreso:
            callback_progreso(total_items, total_items)
        return "SKIPPED"
    # --------------------------------------------------

    logger.info(f"📊 Iniciando generación de Excel completo. Total propiedades a procesar: {total_items}")

    # Listas para las hojas
    data_main = []
    data_constr = []
    data_roles = []   # Separado: Roles Asociados + Su Deuda
    data_deudas = []  # Separado: Solo Deuda Rol Principal
    data_comps = []

    for idx, item in enumerate(lista_datos):
        if cancel_event.is_set():
            logger.warning("🛑 Proceso cancelado por usuario durante la iteración de datos.")
            return False
        # ---------------------------------------------------
        # 1. Actualizar progreso (Llamamos a la función que nos pasaron)
        # ---------------------------------------------------
        if callback_progreso:
            callback_progreso(idx, total_items)
        # ---------------------------------------------------
        
        # --- EXTRACTORES ---
        uid = item.get("ID_Propiedad")
        info_gral = item.get("informacion_general", {})
        avaluo = item.get("avaluo", {})
        transaccion = item.get("transaccion", {})
        carac = item.get("caracteristicas", {})
        info_cbr = item.get("informacion_cbr", {})
        hp_data = item.get("house_pricing", {})
        
        rol_principal = info_gral.get("rol", "S/R")
        logger.debug(f"   [{idx+1}/{total_items}] Procesando ID: {uid} | Rol: {rol_principal}")

        # --- 1. HOJA PRINCIPAL (RESUMEN) ---
        compradores_str = ", ".join(transaccion.get("compradores", []))
        vendedores_str = ", ".join(transaccion.get("vendedores", []))

        # Estado HP
        comps = hp_data.get("comparables", [])
        estado_hp = "Con Resultados"
        if isinstance(comps, str):
            estado_hp = comps 
            num_comps = 0
            logger.debug(f"     ⚠️ Estado HP para {rol_principal}: {estado_hp}")
        else:
            num_comps = len(comps)

        fila_main = {
            "ID Interno": uid,
            "Archivo Origen": item.get("meta_archivo", {}).get("nombre"),
            "link informe": item.get("meta_archivo", {}).get("link_informe"),
            
            # Identificación
            "Rol SII": info_gral.get("rol"),
            "Comuna": info_gral.get("comuna"),
            "Dirección": info_gral.get("direccion"),
            "Propietario": info_gral.get("propietario"),
            
            # Características Globales
            "Tipo Propiedad": carac.get("Tipo"),
            "Destino": carac.get("Destino"),
            "M2 Util": carac.get("M2 Construcción"),
            "M2 Terreno": carac.get("M2 Terreno"),
            
            # Avalúo
            "Avalúo Total": avaluo.get("Avalúo Total"),
            "Avalúo Exento": avaluo.get("Avalúo Exento"),
            "Avalúo Afecto": avaluo.get("Avalúo Afecto"),
            "Contribuciones": avaluo.get("Contribuciones Semestrales"),
            
            # Datos CBR
            "CBR Foja": info_cbr.get("Foja"),
            "CBR Número": info_cbr.get("Número"),
            "CBR Año": info_cbr.get("Año"),
            
            # Transacción
            "Fecha Transacción": transaccion.get("fecha"),
            "Monto Transacción": transaccion.get("monto"),
            "Compradores": compradores_str,
            "Vendedores": vendedores_str,
            
            # Metadata HP
            "Estado Búsqueda HP": estado_hp,
            "Cant. Comparables": num_comps,
            "Latitud Origen": hp_data.get("centro_mapa", {}).get("lat"),
            "Longitud Origen": hp_data.get("centro_mapa", {}).get("lng"),

            # --- NUEVO: TASACIONES ---
            "Tasación Venta CLP": item.get("tasa_vta_clp", 0),
            "Tasación Venta UF": item.get("tasa_vta_uf", "0"),
            "Tasación Arriendo CLP": item.get("tasa_arr_clp", 0),
            "Tasación Arriendo UF": item.get("tasa_arr_uf", "0")
        }
        data_main.append(fila_main)

        # --- 2. HOJA CONSTRUCCIONES ---
        construcciones = item.get("construcciones", [])
        if construcciones:
            logger.debug(f"     🏗️ Agregando {len(construcciones)} construcciones.")
        for c in construcciones:
            data_constr.append({
                "ID Interno (FK)": uid,
                "Rol Propiedad": info_gral.get("rol"),
                "Nro": c.get("nro"),
                "Material": c.get("material"),
                "Calidad": c.get("calidad"),
                "Año": c.get("anio"),
                "M2": c.get("m2"),
                "Destino": c.get("destino")
            })

        # --- PREPARACIÓN: MAPEO DE DEUDAS ---
        mapa_deudas = {}
        for d in item.get("deudas", []):
            key_d = str(d.get("rol", "")).upper().strip()
            mapa_deudas[key_d] = d

        # --- 3. HOJA ROLES ASOCIADOS (MODIFICADO: AHORA INCLUYE DEUDA) ---
        roles_cbr = item.get("roles_cbr", [])
        if roles_cbr:
             logger.debug(f"     📚 Agregando {len(roles_cbr)} roles asociados CBR.")
        
        for r_cbr in roles_cbr:
            rol_asoc_raw = r_cbr.get("rol", "")
            rol_asoc_key = str(rol_asoc_raw).upper().strip()
            
            deuda_obj = mapa_deudas.get(rol_asoc_key, {})
            monto_asoc = deuda_obj.get("monto", 0)
            link_asoc = deuda_obj.get("link_tgr", "No detectado")
            if not link_asoc: link_asoc = "No detectado"

            data_roles.append({
                "ID Interno (FK)": uid,
                "Rol Propiedad": info_gral.get("rol"),
                "Rol Asociado": rol_asoc_raw,
                "Tipo / Ubicación": r_cbr.get("tipo"),
                "Monto Deuda": monto_asoc,  
                "Link Deuda TGR": link_asoc 
            })
        
        # --- 4. HOJA DEUDAS TGR (MODIFICADO: FILTRO SOLO ROL PRINCIPAL) ---
        deudas_list = item.get("deudas", [])
        if deudas_list:
            logger.debug(f"     💰 Procesando deudas (Filtrando solo Rol Principal)...")
        
        for deuda in deudas_list:
            rol_deuda_str = str(deuda.get("rol", ""))
            
            if str(rol_principal) in rol_deuda_str:
                
                link = deuda.get("link_tgr", "")
                if not link: link = "No detectado"
                
                data_deudas.append({
                    "ID Interno (FK)": uid,
                    "Rol Propiedad": info_gral.get("rol"),
                    "Rol Deuda": rol_deuda_str,
                    "Monto Deuda": deuda.get("monto"),
                    "Link informe de deuda TGR": link
                })

        # --- 5. HOJA COMPARABLES ---
        if isinstance(comps, list) and num_comps > 0:
            logger.debug(f"     🏘️ Agregando {num_comps} comparables de mercado.")
            for comp in comps:
                if cancel_event.is_set(): 
                    logger.warning("🛑 Cancelado dentro del loop de comparables.")
                    return False
                data_comps.append({
                    "ID Interno (FK)": uid,
                    "Fuente": comp.get("fuente"),
                    "Rol Origen": info_gral.get("rol"),
                    "Comuna": info_gral.get("comuna"),
                    "Rol Comparable": comp.get("rol"),
                    "Dirección": comp.get("direccion"),
                    "Precio UF": comp.get("precio_uf"),
                    "UF/M2": comp.get("uf_m2"),
                    "Fecha Transacción": comp.get("fecha_transaccion"),
                    "Año Const.": comp.get("anio"),
                    "M2 Útil": comp.get("m2_util"),
                    "M2 Total": comp.get("m2_total"),
                    "Dormitorios": comp.get("dormitorios"),
                    "Baños": comp.get("banios"),
                    "Estacionamientos": comp.get("estacionamientos"),
                    "Bodegas": comp.get("bodegas"),
                    "Distancia (mts)": comp.get("distancia_metros"),
                    "Link Mapa": comp.get("link_maps", ""),
                    "Link Publicacion": comp.get("link_publicacion", "")
                })

    # --- CREACIÓN DE DATAFRAMES ---
    logger.info("💾 Transformando listas a DataFrames...")
    df_main = pd.DataFrame(data_main)
    df_constr = pd.DataFrame(data_constr)
    df_roles = pd.DataFrame(data_roles)
    df_deudas = pd.DataFrame(data_deudas)
    df_comps = pd.DataFrame(data_comps)
    
    logger.debug(f"   📊 [Resumen] Dimensiones: {df_main.shape}")
    logger.debug(f"   📊 [Construcciones] Dimensiones: {df_constr.shape}")
    logger.debug(f"   📊 [Roles Asociados] Dimensiones: {df_roles.shape}")
    logger.debug(f"   📊 [Deudas TGR (Solo Princ.)] Dimensiones: {df_deudas.shape}")
    logger.debug(f"   📊 [Comparables] Dimensiones: {df_comps.shape}")

    try:
        logger.info(f"✍️ Escribiendo archivo físico: {nombre_archivo}")
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            
            # 1. Resumen
            if not df_main.empty:
                logger.debug("   -> Escribiendo hoja 'Resumen General'...")
                df_main.to_excel(writer, sheet_name="Resumen General", index=False)
                _ajustar_columnas(writer, "Resumen General", df_main, cancel_event)
            else:
                logger.warning("   ⚠️ Dataframe 'Resumen General' está vacío.")
            
            # 2. Comparables
            if not df_comps.empty:
                logger.debug("   -> Escribiendo hoja 'Comparables Mercado'...")
                df_comps.to_excel(writer, sheet_name="Comparables Mercado", index=False)
                _ajustar_columnas(writer, "Comparables Mercado", df_comps, cancel_event)
            else:
                logger.debug("   ℹ️ No hay comparables para escribir.")

            # 3. Construcciones
            if not df_constr.empty:
                logger.debug("   -> Escribiendo hoja 'Detalle Construcciones'...")
                df_constr.to_excel(writer, sheet_name="Detalle Construcciones", index=False)
                _ajustar_columnas(writer, "Detalle Construcciones", df_constr, cancel_event)
            
            # 4. Roles Asociados
            if not df_roles.empty:
                logger.debug("   -> Escribiendo hoja 'Roles Asociados'...")
                df_roles.to_excel(writer, sheet_name="Roles Asociados", index=False)
                _ajustar_columnas(writer, "Roles Asociados", df_roles, cancel_event)

            # 5. Deudas TGR
            if not df_deudas.empty:
                logger.debug("   -> Escribiendo hoja 'Deudas TGR'...")
                df_deudas.to_excel(writer, sheet_name="Deudas TGR", index=False)
                _ajustar_columnas(writer, "Deudas TGR", df_deudas, cancel_event)

        logger.success(f"✅ Excel completo generado exitosamente: {nombre_archivo}")
        
        if callback_progreso:
            callback_progreso(total_items, total_items)
        
        return True

    # --- NUEVA SEMÁNTICA: Atrapar errores específicos de escritura en disco ---
    except PermissionError as e:
        mensaje_accionable = "No se pudo generar el reporte. El archivo Excel destino está abierto o bloqueado por otro programa. Ciérrelo e intente nuevamente."
        logger.error(f"❌ Error de permisos en disco: {mensaje_accionable}")
        # En vez de devolver False en silencio, lanzamos el error para que la API lo atrape
        raise Exception(mensaje_accionable)
        
    except Exception as e:
        mensaje_accionable = f"Error estructural al crear el archivo Excel. Es posible que los datos extraídos contengan caracteres inválidos. Detalle técnico: {str(e)}"
        logger.error(f"❌ Error FATAL al guardar Excel completo: {mensaje_accionable}", exc_info=True)
        raise Exception(mensaje_accionable)

def _ajustar_columnas(writer, sheet_name, df, cancel_event):
    """Función auxiliar para auto-ajustar el ancho de columnas (Soporte > 26 cols)"""
    # IMPORTANTE: Importamos la utilidad para letras de columnas (A, B... AA, AB...)
    from openpyxl.utils import get_column_letter 

    logger.debug(f"   🎨 Ajustando formato (Ancho/Links) en: {sheet_name}")
    worksheet = writer.sheets[sheet_name]
    
    for idx, col in enumerate(df.columns):
        if cancel_event.is_set(): return
        
        # --- CORRECCIÓN BUG '[' ---
        # Usamos get_column_letter(idx + 1) en lugar de chr(65 + idx)
        # Esto convierte el índice 26 en 'AA' correctamente.
        col_letter = get_column_letter(idx + 1)
        
        try:
            # Calculamos ancho basado en el contenido
            max_len_data = df[col].astype(str).map(len).max() if not df[col].empty else 0
            max_len = max(max_len_data, len(str(col))) + 2
            max_len = min(max_len, 60) # Tope máximo
            
            worksheet.column_dimensions[col_letter].width = max_len
        except Exception as e:
            logger.debug(f"      ⚠️ No se pudo ajustar ancho col {col}: {e}")
            pass
        
        # Detección de Links para formato azul y clickable
        if "Link" in str(col):
            count_links = 0
            for row_idx in range(2, len(df) + 2):
                cell = worksheet[f"{col_letter}{row_idx}"] # Usamos la letra corregida
                val = cell.value
                
                # Si el valor empieza con http, lo hacemos clickable y azul
                if val and str(val).startswith("http"):
                    cell.hyperlink = val
                    cell.style = "Hyperlink" 
                    count_links += 1
            
            if count_links > 0:
                logger.debug(f"      🔗 {count_links} hipervínculos formateados en columna {col}")